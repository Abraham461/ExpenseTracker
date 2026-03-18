from datetime import date, datetime, timedelta
from decimal import Decimal
import csv
from io import BytesIO, StringIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.db.models import Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render

from .budget import get_budget_context
from .forms import BudgetForm, CategoryForm, ExpenseForm, SignUpForm
from .insights import get_monthly_insights
from .models import Category, Expense, ExpenseNotification, UserProfile

DEFAULT_CATEGORY_NAMES = [
    'Food',
    'Transport',
    'Shopping',
    'Bills',
    'Entertainment',
    'Health',
    'Education',
    'Travel',
    'Personal Care',
    'Other',
]


def _ensure_default_categories(user):
    existing = set(user.categories.values_list('name', flat=True))
    missing = [name for name in DEFAULT_CATEGORY_NAMES if name not in existing]
    if missing:
        Category.objects.bulk_create(
            [Category(user=user, name=name) for name in missing],
            ignore_conflicts=True,
        )


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            _ensure_default_categories(user)
            login(request, user)
            request.session['show_budget_prompt'] = True
            return redirect('dashboard')
    else:
        form = SignUpForm()
    return render(request, 'registration/signup.html', {'form': form})


@login_required
def dashboard(request):
    today = date.today()
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    _ensure_default_categories(request.user)
    show_budget_prompt = request.session.pop('show_budget_prompt', False)

    expenses = request.user.expenses.select_related('category')
    month_start = today.replace(day=1)
    month_total = expenses.filter(spent_on__gte=month_start, spent_on__lte=today).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    today_total = expenses.filter(spent_on=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    daily_limit = None
    base_daily_limit = None
    month_allowed = None
    remaining_today = None
    remaining_month = None
    period_start = None
    period_budget = None
    period_total = None
    days_in_period = None
    days_elapsed = None
    month_end = None
    month_start = None
    days_in_month = None
    budget = get_budget_context(request.user, today=today, profile=profile)
    if budget['has_budget']:
        base_daily_limit = budget['base_daily_limit']
        month_allowed = budget['allowed_to_date']
        daily_limit = budget['daily_limit']
        remaining_today = budget['remaining_today']
        remaining_month = budget['remaining_period']
        month_start = budget['month_start']
        period_start = budget['period_start']
        period_budget = budget['period_budget']
        period_total = budget['period_total']
        days_in_month = budget['days_in_month']
        days_in_period = budget['days_in_period']
        days_elapsed = budget['days_elapsed']
        month_end = budget['month_end']

    recent_expenses = expenses.order_by('-spent_on', '-created_at')[:6]
    top_categories = (
        expenses.filter(spent_on__gte=month_start, spent_on__lte=today)
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')[:5]
    )

    if not budget['has_budget']:
        show_budget_prompt = True

    budget_status = None
    budget_message = None
    budget_percent = None
    if budget['has_budget'] and base_daily_limit and base_daily_limit > 0:
        usage_percent = (today_total / base_daily_limit * Decimal('100')).quantize(Decimal('0.1'))
        usage_percent_value = float(usage_percent)
        budget_percent = min(100.0, usage_percent_value)
        if today_total > base_daily_limit:
            budget_status = 'danger'
            overspend = (today_total - base_daily_limit).quantize(Decimal('0.01'))
            budget_message = (
                f'You exceeded today\'s budget by {settings.CURRENCY_PREFIX}{overspend}.'
            )
        else:
            budget_status = 'success' if usage_percent_value <= 80 else 'warning'
            budget_message = f'You have used {usage_percent_value:.0f}% of today\'s budget.'

    expense_form = ExpenseForm(user=request.user, initial={'spent_on': today})

    context = {
        'profile': profile,
        'daily_limit': daily_limit,
        'base_daily_limit': base_daily_limit,
        'month_allowed': month_allowed,
        'month_total': month_total,
        'today_total': today_total,
        'remaining_today': remaining_today,
        'remaining_month': remaining_month,
        'period_start': period_start,
        'month_start': month_start,
        'period_budget': period_budget,
        'days_in_month': days_in_month,
        'period_total': period_total,
        'days_in_period': days_in_period,
        'days_elapsed': days_elapsed,
        'month_end': month_end,
        'show_budget_prompt': show_budget_prompt,
        'budget_form': BudgetForm(instance=profile),
        'expense_form': expense_form,
        'recent_expenses': recent_expenses,
        'top_categories': top_categories,
        'budget_status': budget_status,
        'budget_message': budget_message,
        'budget_percent': budget_percent,
    }
    return render(request, 'subscriptions/dashboard.html', context)


@login_required
def expense_list(request):
    expenses = request.user.expenses.select_related('category')
    return render(
        request,
        'subscriptions/subscription_list.html',
        {'expenses': expenses},
    )


@login_required
def expense_create(request):
    _ensure_default_categories(request.user)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, user=request.user)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            _maybe_send_overspend_alert(request.user, date.today())
            messages.success(request, 'Expense added successfully.')
            next_url = request.POST.get('next')
            return redirect(next_url or 'expense_list')
    else:
        form = ExpenseForm(user=request.user, initial={'spent_on': date.today()})
    return render(request, 'subscriptions/subscription_form.html', {'form': form, 'title': 'Add Expense'})


@login_required
def expense_update(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if expense.user != request.user:
        return HttpResponseForbidden()

    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense, user=request.user)
        if form.is_valid():
            form.save()
            _maybe_send_overspend_alert(request.user, date.today())
            messages.success(request, 'Expense updated.')
            return redirect('expense_list')
    else:
        form = ExpenseForm(instance=expense, user=request.user)
    return render(request, 'subscriptions/subscription_form.html', {'form': form, 'title': 'Edit Expense'})


@login_required
def expense_delete(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if expense.user != request.user:
        return HttpResponseForbidden()
    if request.method == 'POST':
        expense.delete()
        messages.success(request, 'Expense deleted.')
        return redirect('expense_list')
    return render(request, 'subscriptions/subscription_delete.html', {'expense': expense})


@login_required
def category_list(request):
    _ensure_default_categories(request.user)
    return render(request, 'subscriptions/category_list.html', {'categories': request.user.categories.all()})


@login_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, user=request.user)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            return redirect('category_list')
    else:
        form = CategoryForm(user=request.user)
    return render(request, 'subscriptions/category_form.html', {'form': form})


@login_required
def analytics(request):
    expenses = request.user.expenses.select_related('category')
    totals = (
        expenses.values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    labels = [item['category__name'] or 'Uncategorized' for item in totals]
    amounts = [float(item['total']) for item in totals]

    today = date.today()
    start_date = today - timedelta(days=6)
    weekly_totals = (
        expenses.filter(spent_on__gte=start_date, spent_on__lte=today)
        .values('spent_on')
        .annotate(total=Sum('amount'))
    )
    weekly_map = {item['spent_on']: float(item['total']) for item in weekly_totals}
    weekly_labels = []
    weekly_amounts = []
    for offset in range(7):
        day = start_date + timedelta(days=offset)
        weekly_labels.append(day.strftime('%a'))
        weekly_amounts.append(weekly_map.get(day, 0.0))

    return render(
        request,
        'subscriptions/analytics.html',
        {
            'labels': labels,
            'amounts': amounts,
            'weekly_labels': weekly_labels,
            'weekly_amounts': weekly_amounts,
        },
    )


@login_required
def insights(request):
    month_param = (request.GET.get('month') or '').strip()
    year = None
    month = None
    if month_param:
        try:
            parts = month_param.split('-')
            if len(parts) == 2:
                year = int(parts[0])
                month = int(parts[1])
                if month < 1 or month > 12:
                    raise ValueError
        except ValueError:
            year = None
            month = None

    data = get_monthly_insights(request.user, year=year, month=month)
    selected_month = f"{data['month_start']:%Y-%m}"

    context = {
        **data,
        'month_label': data['month_start'].strftime('%B %Y'),
        'selected_month': selected_month,
    }
    return render(request, 'subscriptions/insights.html', context)


@login_required
def export_expenses(request, export_format):
    expenses = list(request.user.expenses.select_related('category').order_by('-spent_on', '-created_at'))
    headers = ['ID', 'Date', 'Amount', 'Currency', 'Category', 'Reason', 'Note', 'Created At']
    if export_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for expense in expenses:
            writer.writerow([
                expense.id,
                expense.spent_on,
                expense.amount,
                settings.CURRENCY_SYMBOL,
                expense.category.name if expense.category else 'Uncategorized',
                expense.reason,
                expense.note,
                expense.created_at.isoformat(),
            ])
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="expenses.csv"'
        return response

    if export_format == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Expenses'
        sheet.append(headers)
        for expense in expenses:
            sheet.append([
                expense.id,
                expense.spent_on.isoformat(),
                float(expense.amount),
                settings.CURRENCY_SYMBOL,
                expense.category.name if expense.category else 'Uncategorized',
                expense.reason,
                expense.note,
                expense.created_at.isoformat(),
            ])
        categories = list(request.user.categories.order_by('name').values_list('name', flat=True))
        if not categories:
            categories = ['Uncategorized']
        category_sheet = workbook.create_sheet('Categories')
        for idx, name in enumerate(categories, start=1):
            category_sheet.cell(row=idx, column=1, value=name)
        formula = f'=Categories!$A$1:$A${len(categories)}'
        validation = DataValidation(type='list', formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        max_rows = max(len(expenses) + 1, 1000)
        validation.add(f'E2:E{max_rows}')
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="expenses.xlsx"'
        return response

    return HttpResponse(status=404)


@login_required
def export_expense_template(request):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    _ensure_default_categories(request.user)
    categories = list(request.user.categories.order_by('name').values_list('name', flat=True))
    if not categories:
        categories = ['Uncategorized']

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Expenses'
    headers = ['Date', 'Amount', 'Category', 'Reason', 'Note']
    sheet.append(headers)
    sheet.append([date.today().isoformat(), '', categories[0], '', ''])
    sheet.freeze_panes = 'A2'

    category_sheet = workbook.create_sheet('Categories')
    for idx, name in enumerate(categories, start=1):
        category_sheet.cell(row=idx, column=1, value=name)

    formula = f'=Categories!$A$1:$A${len(categories)}'
    validation = DataValidation(type='list', formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add('C2:C1000')

    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="expense_import_template.xlsx"'
    return response


@login_required
def import_expenses(request):
    if request.method == 'POST' and request.FILES.get('file'):
        upload = request.FILES['file']
        if not upload.name.lower().endswith('.xlsx'):
            messages.error(request, 'Please upload an .xlsx file.')
            return redirect('import_expenses')

        from openpyxl import load_workbook

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            messages.error(request, 'Unable to read the Excel file.')
            return redirect('import_expenses')

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            messages.error(request, 'The spreadsheet is empty.')
            return redirect('import_expenses')

        def normalize(value):
            return str(value or '').strip().lower()

        header_row = rows[0]
        header_map = {normalize(name): idx for idx, name in enumerate(header_row)}

        def find_header(candidates):
            for name in candidates:
                if name in header_map:
                    return header_map[name]
            return None

        date_idx = find_header(['date', 'spent on', 'spent_on'])
        amount_idx = find_header(['amount', 'cost', 'value'])
        category_idx = find_header(['category', 'type'])
        reason_idx = find_header(['reason', 'description'])
        note_idx = find_header(['note', 'notes', 'memo'])

        if date_idx is None or amount_idx is None:
            messages.error(request, 'Missing required columns: Date and Amount.')
            return redirect('import_expenses')

        _ensure_default_categories(request.user)
        category_lookup = {
            category.name.lower(): category for category in request.user.categories.all()
        }

        created = 0
        skipped = 0
        duplicates = 0

        for row in rows[1:]:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            date_cell = row[date_idx] if date_idx < len(row) else None
            amount_cell = row[amount_idx] if amount_idx < len(row) else None
            if amount_cell is None:
                skipped += 1
                continue

            spent_on = None
            if isinstance(date_cell, datetime):
                spent_on = date_cell.date()
            elif isinstance(date_cell, date):
                spent_on = date_cell
            elif isinstance(date_cell, str):
                try:
                    spent_on = datetime.fromisoformat(date_cell.strip()).date()
                except ValueError:
                    spent_on = None

            if spent_on is None:
                skipped += 1
                continue

            try:
                amount = Decimal(str(amount_cell))
            except Exception:
                skipped += 1
                continue

            category_name = ''
            if category_idx is not None and category_idx < len(row):
                category_name = str(row[category_idx] or '').strip()
            category = None
            if category_name:
                key = category_name.lower()
                category = category_lookup.get(key)
                if category is None:
                    category = Category.objects.create(user=request.user, name=category_name)
                    category_lookup[key] = category

            reason = ''
            if reason_idx is not None and reason_idx < len(row):
                reason = str(row[reason_idx] or '').strip()

            note = ''
            if note_idx is not None and note_idx < len(row):
                note = str(row[note_idx] or '').strip()

            if Expense.objects.filter(
                user=request.user,
                spent_on=spent_on,
                amount=amount,
                category=category,
                reason=reason,
                note=note,
            ).exists():
                duplicates += 1
                continue

            Expense.objects.create(
                user=request.user,
                amount=amount,
                category=category,
                reason=reason,
                spent_on=spent_on,
                note=note,
            )
            created += 1

        messages.success(
            request,
            f'Imported {created} expenses. Skipped {skipped} rows ({duplicates} duplicates).'
        )
        return redirect('expense_list')

    return render(request, 'subscriptions/expense_import.html')


@login_required
def income_settings(request):
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Budget updated.')
            return redirect('dashboard')
    else:
        form = BudgetForm(instance=profile)
    return render(request, 'subscriptions/usage_form.html', {'form': form, 'subscription': None, 'title': 'Budget Settings'})


@login_required
def run_reminders(request):
    today = date.today()
    daily_sent = _maybe_send_daily_reminder(request.user, today)
    overspend_sent = _maybe_send_overspend_alert(request.user, today)
    monthly_sent = _maybe_send_monthly_insight(request.user, today)

    status = []
    if daily_sent:
        status.append('daily reminder sent')
    if overspend_sent:
        status.append('overspend alert sent')
    if monthly_sent:
        status.append('monthly insight sent')
    if not status:
        status.append('no notifications needed')
    messages.success(request, ', '.join(status).capitalize() + '.')
    return redirect('dashboard')


def _maybe_send_daily_reminder(user, today):
    if not user.email:
        return False
    profile, _ = UserProfile.objects.get_or_create(user=user)
    if not profile.daily_email_reminders:
        return False
    if Expense.objects.filter(user=user, spent_on=today).exists():
        return False
    if ExpenseNotification.objects.filter(
        user=user, notification_type=ExpenseNotification.TYPE_DAILY, sent_at__date=today
    ).exists():
        return False
    message = render_to_string(
        'emails/daily_reminder.txt',
        {'user': user, 'date': today, 'currency_prefix': settings.CURRENCY_PREFIX},
    ).strip()
    send_mail(
        subject='Daily Expense Reminder',
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=settings.EMAIL_FAIL_SILENTLY,
    )
    ExpenseNotification.objects.create(
        user=user, notification_type=ExpenseNotification.TYPE_DAILY, message=message
    )
    return True


def _maybe_send_overspend_alert(user, today):
    if not user.email:
        return False
    profile, _ = UserProfile.objects.get_or_create(user=user)
    budget = get_budget_context(user, today=today, profile=profile)
    if not budget['has_budget']:
        return False
    daily_limit = budget['base_daily_limit']
    today_total = budget['today_total']
    if daily_limit is None or daily_limit <= 0:
        return False
    if today_total <= daily_limit:
        return False
    if ExpenseNotification.objects.filter(
        user=user, notification_type=ExpenseNotification.TYPE_OVERSPEND, sent_at__date=today
    ).exists():
        return False
    overspend = (today_total - daily_limit).quantize(Decimal('0.01'))
    message = render_to_string(
        'emails/overspend_alert.txt',
        {
            'user': user,
            'date': today,
            'currency_prefix': settings.CURRENCY_PREFIX,
            'overspend': overspend,
        },
    ).strip()
    send_mail(
        subject='Overspending Alert',
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=settings.EMAIL_FAIL_SILENTLY,
    )
    ExpenseNotification.objects.create(
        user=user, notification_type=ExpenseNotification.TYPE_OVERSPEND, message=message
    )
    return True


def _maybe_send_monthly_insight(user, today):
    if not user.email:
        return False
    if ExpenseNotification.objects.filter(
        user=user,
        notification_type=ExpenseNotification.TYPE_MONTHLY,
        sent_at__year=today.year,
        sent_at__month=today.month,
    ).exists():
        return False
    month_end = today.replace(day=1) - timedelta(days=1)
    data = get_monthly_insights(user, year=month_end.year, month=month_end.month)
    message = render_to_string(
        'emails/monthly_insight.txt',
        {
            'user': user,
            'month_label': data['month_start'].strftime('%B %Y'),
            'month_total': data['month_total'],
            'prev_month_total': data['prev_month_total'],
            'change_amount': data['change_amount'],
            'change_percent': data['change_percent'],
            'days_logged': data['days_logged'],
            'has_expenses': data['has_expenses'],
            'common_reasons': data['common_reasons'],
            'currency_prefix': settings.CURRENCY_PREFIX,
        },
    ).strip()
    send_mail(
        subject=f'Monthly Insight: {data["month_start"]:%B %Y}',
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=settings.EMAIL_FAIL_SILENTLY,
    )
    ExpenseNotification.objects.create(
        user=user, notification_type=ExpenseNotification.TYPE_MONTHLY, message=message
    )
    return True
